#!/usr/bin/env python
from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data" / "functional_imaging" / "dpm_lateralization"
DEFAULT_OLD_XLSX = DATA_DIR / "2026-05-21_dpm_lateralization_functional_imaging.xlsx"
DEFAULT_NEW_XLSX = DATA_DIR / "2026-06-03_dpm_lateralization_functional_imaging.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "outputs" / "dpm_functional_imaging_lateralization_20260603"


@dataclass(frozen=True)
class ResponseWindow:
    sheet: str
    pulse: int
    start_excel_row: int
    end_excel_row: int
    start_s: float
    end_s: float
    n_points: int
    formula: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Reanalyse DPM left-right functional-imaging Excel files and compare old/new/merged batches."
    )
    parser.add_argument("--old-xlsx", type=Path, default=DEFAULT_OLD_XLSX)
    parser.add_argument("--new-xlsx", type=Path, default=DEFAULT_NEW_XLSX)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def parse_pulse(sheet_name: str) -> int:
    if "1-pulse" in sheet_name:
        return 1
    match = re.search(r"(\d+)\s*pulses?", sheet_name)
    if not match:
        raise ValueError(f"Cannot parse pulse count from sheet name: {sheet_name!r}")
    return int(match.group(1))


def parse_sum_formula(formula: str) -> tuple[int, int]:
    match = re.search(r"SUM\([A-Z]+(\d+):[A-Z]+(\d+)\)", str(formula))
    if not match:
        raise ValueError(f"Cannot parse response-window formula: {formula!r}")
    return int(match.group(1)), int(match.group(2))


def load_response_windows(xlsx_path: Path, sheets: list[str]) -> dict[str, ResponseWindow]:
    workbook = load_workbook(xlsx_path, data_only=False, read_only=True)
    windows: dict[str, ResponseWindow] = {}
    for sheet in sheets:
        ws = workbook[sheet]
        formula = str(ws["B304"].value)
        start_row, end_row = parse_sum_formula(formula)
        windows[sheet] = ResponseWindow(
            sheet=sheet,
            pulse=parse_pulse(sheet),
            start_excel_row=start_row,
            end_excel_row=end_row,
            start_s=float(ws[f"A{start_row}"].value),
            end_s=float(ws[f"A{end_row}"].value),
            n_points=end_row - start_row + 1,
            formula=formula,
        )
    workbook.close()
    return windows


def iter_fly_side_columns(raw: pd.DataFrame) -> list[tuple[str, str, int]]:
    columns: list[tuple[str, str, int]] = []
    current_fly: str | None = None
    for col in range(1, raw.shape[1]):
        fly_cell = raw.iloc[0, col]
        if isinstance(fly_cell, str) and fly_cell.strip().lower().startswith("fly"):
            current_fly = fly_cell.strip().lower()
        side = str(raw.iloc[1, col]).strip().lower()
        if current_fly is None or side not in {"left", "right"}:
            continue
        values = pd.to_numeric(raw.iloc[2:302, col], errors="coerce")
        if values.notna().sum() < 250:
            continue
        columns.append((current_fly, side, col))
    return columns


def load_excel_dataset(xlsx_path: Path, dataset: str, fly_prefix: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    excel = pd.ExcelFile(xlsx_path)
    windows = load_response_windows(xlsx_path, excel.sheet_names)
    time_rows: list[dict] = []
    metric_rows: list[dict] = []
    window_rows: list[dict] = []

    for sheet in excel.sheet_names:
        raw = pd.read_excel(xlsx_path, sheet_name=sheet, header=None)
        window = windows[sheet]
        window_rows.append({"dataset": dataset, **window.__dict__})

        time_df = pd.DataFrame(
            {
                "excel_row": np.arange(3, 303),
                "time_s": pd.to_numeric(raw.iloc[2:302, 0], errors="coerce").to_numpy(float),
            }
        )
        baseline_mask = time_df["excel_row"] < window.start_excel_row
        stim_mask = (time_df["excel_row"] >= window.start_excel_row) & (
            time_df["excel_row"] <= window.end_excel_row
        )
        post_mask = time_df["excel_row"] >= window.start_excel_row
        post_30_60_mask = (time_df["time_s"] >= 30.0) & (time_df["time_s"] < 60.0)
        post_60_120_mask = (time_df["time_s"] >= 60.0) & (time_df["time_s"] < 120.0)

        for raw_fly, side, col in iter_fly_side_columns(raw):
            values = pd.to_numeric(raw.iloc[2:302, col], errors="coerce").to_numpy(float)
            fly = f"{fly_prefix}_{raw_fly}"
            side_df = time_df.copy()
            side_df["dataset"] = dataset
            side_df["sheet"] = sheet
            side_df["pulse"] = window.pulse
            side_df["fly"] = fly
            side_df["raw_fly"] = raw_fly
            side_df["side"] = side
            side_df["value"] = values
            time_rows.extend(side_df.to_dict("records"))

            baseline = float(np.nanmean(values[baseline_mask.to_numpy()]))
            stim_values = values[stim_mask.to_numpy()]
            stim_bc = stim_values - baseline
            post_values = values[post_mask.to_numpy()]
            post_bc = post_values - baseline
            row_stim = pd.to_numeric(pd.Series([raw.iloc[303, col]]), errors="coerce").iloc[0]
            row_peak = pd.to_numeric(pd.Series([raw.iloc[305, col]]), errors="coerce").iloc[0]
            metric_rows.append(
                {
                    "dataset": dataset,
                    "sheet": sheet,
                    "pulse": window.pulse,
                    "fly": fly,
                    "raw_fly": raw_fly,
                    "side": side,
                    "baseline_mean_0_to_prestim": baseline,
                    "excel_stim_sum": float(row_stim) if np.isfinite(row_stim) else float(np.nansum(stim_values)),
                    "excel_peak_full": float(row_peak) if np.isfinite(row_peak) else float(np.nanmax(values)),
                    "stim_window_mean_raw": float(np.nanmean(stim_values)),
                    "stim_window_sum_from_formula_rows": float(np.nansum(stim_values)),
                    "stim_window_mean_bc": float(np.nanmean(stim_bc)),
                    "stim_window_sum_bc": float(np.nansum(stim_bc)),
                    "stim_window_peak_bc": float(np.nanmax(stim_bc)),
                    "post_peak_bc": float(np.nanmax(post_bc)),
                    "post_30_60_mean_bc": float(np.nanmean(values[post_30_60_mask.to_numpy()] - baseline)),
                    "post_60_120_mean_bc": float(np.nanmean(values[post_60_120_mask.to_numpy()] - baseline)),
                }
            )

    return pd.DataFrame(time_rows), pd.DataFrame(metric_rows), pd.DataFrame(window_rows)


def paired_differences(metrics: pd.DataFrame, metric: str) -> pd.DataFrame:
    wide = metrics.pivot_table(index=["dataset", "fly", "pulse"], columns="side", values=metric, aggfunc="first")
    wide = wide.dropna(subset=["left", "right"])
    wide["right_minus_left"] = wide["right"] - wide["left"]
    return wide.reset_index()[["dataset", "fly", "pulse", "right_minus_left"]]


def paired_summary(values: pd.Series | np.ndarray) -> dict:
    x = np.asarray(pd.Series(values).dropna(), dtype=float)
    n = int(x.size)
    positives = int(np.sum(x > 0))
    negatives = int(np.sum(x < 0))
    sign_n = positives + negatives
    if n < 2:
        return {
            "n": n,
            "mean_right_minus_left": np.nan,
            "median_right_minus_left": np.nan,
            "sd_right_minus_left": np.nan,
            "cohen_dz": np.nan,
            "paired_t_p_two_sided": np.nan,
            "paired_t_p_greater_right": np.nan,
            "wilcoxon_p_two_sided": np.nan,
            "sign_test_p_two_sided": np.nan,
            "positive_fly_count": positives,
            "negative_fly_count": negatives,
        }
    mean = float(np.mean(x))
    sd = float(np.std(x, ddof=1))
    try:
        wilcoxon_p = float(stats.wilcoxon(x, zero_method="wilcox", alternative="two-sided").pvalue)
    except ValueError:
        wilcoxon_p = np.nan
    return {
        "n": n,
        "mean_right_minus_left": mean,
        "median_right_minus_left": float(np.median(x)),
        "sd_right_minus_left": sd,
        "cohen_dz": mean / sd if sd > 0 else np.nan,
        "paired_t_p_two_sided": float(stats.ttest_1samp(x, 0.0, alternative="two-sided").pvalue),
        "paired_t_p_greater_right": float(stats.ttest_1samp(x, 0.0, alternative="greater").pvalue),
        "wilcoxon_p_two_sided": wilcoxon_p,
        "sign_test_p_two_sided": float(stats.binomtest(positives, sign_n, 0.5, alternative="two-sided").pvalue)
        if sign_n
        else np.nan,
        "positive_fly_count": positives,
        "negative_fly_count": negatives,
    }


def build_test_tables(metrics_by_label: dict[str, pd.DataFrame], metric_names: list[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    per_pulse_rows: list[dict] = []
    global_rows: list[dict] = []
    pulses = [1, 50, 100, 200]
    log_pulses = np.log10(np.array(pulses, dtype=float))

    for dataset_label, metrics in metrics_by_label.items():
        for metric in metric_names:
            diff = paired_differences(metrics, metric)
            for pulse in pulses:
                values = diff.loc[diff["pulse"].eq(pulse), "right_minus_left"]
                per_pulse_rows.append(
                    {
                        "dataset_group": dataset_label,
                        "metric": metric,
                        "contrast": f"pulse_{pulse}",
                        "pulse": pulse,
                        **paired_summary(values),
                    }
                )

            wide = diff.pivot_table(index=["dataset", "fly"], columns="pulse", values="right_minus_left")
            wide = wide[pulses].dropna()
            contrasts = {
                "all_pulses_mean": wide[pulses].mean(axis=1),
                "high_pulses_50_100_200_mean": wide[[50, 100, 200]].mean(axis=1),
                "high_pulses_mean_minus_1_pulse": wide[[50, 100, 200]].mean(axis=1) - wide[1],
            }
            slopes = []
            for _, row in wide.iterrows():
                slopes.append(np.polyfit(log_pulses, row.to_numpy(float), 1)[0])
            contrasts["dose_slope_log10_pulse"] = pd.Series(slopes)
            for contrast, values in contrasts.items():
                global_rows.append(
                    {
                        "dataset_group": dataset_label,
                        "metric": metric,
                        "contrast": contrast,
                        **paired_summary(values),
                    }
                )

    per_pulse = pd.DataFrame(per_pulse_rows)
    global_tests = pd.DataFrame(global_rows)
    return per_pulse, global_tests


def fmt_p(value: float) -> str:
    if not np.isfinite(value):
        return "NA"
    if value < 0.001:
        return f"{value:.2e}"
    return f"{value:.4f}"


def result_row(table: pd.DataFrame, dataset_group: str, metric: str, contrast: str) -> pd.Series:
    match = table[
        table["dataset_group"].eq(dataset_group)
        & table["metric"].eq(metric)
        & table["contrast"].eq(contrast)
    ]
    if match.empty:
        raise KeyError((dataset_group, metric, contrast))
    return match.iloc[0]


def plot_summary(per_pulse: pd.DataFrame, global_tests: pd.DataFrame, figures_dir: Path) -> Path:
    figures_dir.mkdir(parents=True, exist_ok=True)
    dataset_order = ["old_17", "new_12", "combined_29"]
    colors = {"old_17": "#4C78A8", "new_12": "#F58518", "combined_29": "#54A24B"}
    pulse_order = [1, 50, 100, 200]

    fig, axes = plt.subplots(2, 2, figsize=(12, 8), constrained_layout=True)
    metric_titles = [
        ("excel_stim_sum", "Stimulus-window sum"),
        ("excel_peak_full", "Full-trace peak"),
    ]
    for ax, (metric, title) in zip(axes[0], metric_titles):
        for offset, dataset in zip([-0.18, 0.0, 0.18], dataset_order):
            rows = per_pulse[
                per_pulse["dataset_group"].eq(dataset)
                & per_pulse["metric"].eq(metric)
                & per_pulse["pulse"].isin(pulse_order)
            ].sort_values("pulse")
            x = np.arange(len(pulse_order), dtype=float) + offset
            y = rows["mean_right_minus_left"].to_numpy(float)
            sem = rows["sd_right_minus_left"].to_numpy(float) / np.sqrt(rows["n"].to_numpy(float))
            ax.errorbar(x, y, yerr=sem, fmt="o-", color=colors[dataset], label=dataset, capsize=3, lw=1.4)
        ax.axhline(0, color="#333333", lw=0.8)
        ax.set_xticks(np.arange(len(pulse_order)), [str(p) for p in pulse_order])
        ax.set_xlabel("Pulse count")
        ax.set_ylabel("Mean right - left")
        ax.set_title(title)
        ax.spines[["top", "right"]].set_visible(False)
    axes[0, 0].legend(frameon=False)

    contrast_order = ["all_pulses_mean", "high_pulses_mean_minus_1_pulse", "dose_slope_log10_pulse"]
    contrast_labels = ["all mean", "high - 1", "log-dose slope"]
    for ax, (metric, title) in zip(axes[1], metric_titles):
        width = 0.24
        for idx, dataset in enumerate(dataset_order):
            rows = global_tests[
                global_tests["dataset_group"].eq(dataset)
                & global_tests["metric"].eq(metric)
                & global_tests["contrast"].isin(contrast_order)
            ].set_index("contrast").loc[contrast_order]
            y = -np.log10(rows["paired_t_p_two_sided"].astype(float).clip(lower=1e-300))
            x = np.arange(len(contrast_order), dtype=float) + (idx - 1) * width
            ax.bar(x, y, width=width, color=colors[dataset], label=dataset)
        ax.axhline(-math.log10(0.05), color="#C0392B", lw=1, ls="--")
        ax.set_xticks(np.arange(len(contrast_order)), contrast_labels, rotation=20, ha="right")
        ax.set_ylabel("-log10 paired t p")
        ax.set_title(f"{title}: global contrasts")
        ax.spines[["top", "right"]].set_visible(False)

    out = figures_dir / "Fig_dpm_functional_imaging_lateralization_20260603.png"
    fig.savefig(out, dpi=240)
    plt.close(fig)
    return out


def fly_level_dose_slopes(metrics: pd.DataFrame, metric: str) -> pd.DataFrame:
    pulses = [1, 50, 100, 200]
    log_pulses = np.log10(np.array(pulses, dtype=float))
    wide = metrics.pivot_table(index=["dataset", "fly", "side"], columns="pulse", values=metric, aggfunc="first")
    wide = wide[pulses].dropna()
    rows: list[dict[str, object]] = []
    for (dataset, fly, side), row in wide.iterrows():
        rows.append(
            {
                "dataset": dataset,
                "fly": fly,
                "side": side,
                "metric": metric,
                "dose_slope_log10_pulse": float(np.polyfit(log_pulses, row.to_numpy(float), 1)[0]),
                "high_pulses_mean_minus_1_pulse": float(row[[50, 100, 200]].mean() - row[1]),
            }
        )
    return pd.DataFrame.from_records(rows)


def plot_paired_primary_endpoint(all_metrics: pd.DataFrame, figures_dir: Path) -> Path:
    figures_dir.mkdir(parents=True, exist_ok=True)
    slopes = fly_level_dose_slopes(all_metrics, "excel_peak_full")
    wide = slopes.pivot_table(
        index=["dataset", "fly"],
        columns="side",
        values="dose_slope_log10_pulse",
        aggfunc="first",
    ).dropna(subset=["left", "right"]).reset_index()
    dataset_order = ["2026-05-21_old", "2026-06-03_new"]
    labels = {"2026-05-21_old": "old batch (n=17)", "2026-06-03_new": "new batch (n=12)"}
    colors = {"2026-05-21_old": "#4C78A8", "2026-06-03_new": "#F58518"}

    fig, axes = plt.subplots(1, 3, figsize=(11.5, 4.0), constrained_layout=True)
    for ax, dataset in zip(axes[:2], dataset_order):
        sub = wide[wide["dataset"].eq(dataset)].copy()
        for idx, row in sub.iterrows():
            ax.plot([0, 1], [row["left"], row["right"]], color=colors[dataset], alpha=0.45, lw=1.1)
            ax.scatter([0, 1], [row["left"], row["right"]], color=colors[dataset], s=18, alpha=0.8)
        diff = sub["right"] - sub["left"]
        summary = paired_summary(diff)
        ax.axhline(0, color="#333333", lw=0.7)
        ax.set_xticks([0, 1], ["left", "right"])
        ax.set_ylabel("peak dose slope")
        ax.set_title(f"{labels[dataset]}\nR>L {int(summary['positive_fly_count'])}/{int(summary['n'])}, p={fmt_p(summary['paired_t_p_two_sided'])}")
        ax.spines[["top", "right"]].set_visible(False)

    combined = wide.copy()
    combined["right_minus_left"] = combined["right"] - combined["left"]
    bins = np.linspace(
        min(-0.02, float(combined["right_minus_left"].min()) * 1.05),
        max(0.02, float(combined["right_minus_left"].max()) * 1.05),
        13,
    )
    for dataset in dataset_order:
        sub = combined[combined["dataset"].eq(dataset)]
        axes[2].hist(
            sub["right_minus_left"],
            bins=bins,
            alpha=0.65,
            label=labels[dataset].split()[0],
            color=colors[dataset],
            edgecolor="white",
        )
    summary = paired_summary(combined["right_minus_left"])
    axes[2].axvline(0, color="#333333", lw=0.8)
    axes[2].axvline(combined["right_minus_left"].mean(), color="#C0392B", lw=1.5)
    axes[2].set_xlabel("right - left peak dose slope")
    axes[2].set_ylabel("fly count")
    axes[2].set_title(
        f"combined n={int(summary['n'])}\nR>L {int(summary['positive_fly_count'])}/{int(summary['n'])}, p={fmt_p(summary['paired_t_p_two_sided'])}"
    )
    axes[2].legend(frameon=False)
    axes[2].spines[["top", "right"]].set_visible(False)
    fig.suptitle("DPM functional imaging primary endpoint: fly-level paired peak dose slope", fontsize=12)
    out = figures_dir / "Fig_dpm_paired_peak_dose_slope_primary_endpoint.png"
    fig.savefig(out, dpi=240)
    plt.close(fig)
    return out


def write_report(
    output_dir: Path,
    manifest: pd.DataFrame,
    side_metrics: pd.DataFrame,
    per_pulse: pd.DataFrame,
    global_tests: pd.DataFrame,
    figure_path: Path,
    paired_figure_path: Path,
) -> Path:
    old_stim_slope = result_row(global_tests, "old_17", "excel_stim_sum", "dose_slope_log10_pulse")
    new_stim_slope = result_row(global_tests, "new_12", "excel_stim_sum", "dose_slope_log10_pulse")
    comb_stim_slope = result_row(global_tests, "combined_29", "excel_stim_sum", "dose_slope_log10_pulse")
    old_peak_slope = result_row(global_tests, "old_17", "excel_peak_full", "dose_slope_log10_pulse")
    new_peak_slope = result_row(global_tests, "new_12", "excel_peak_full", "dose_slope_log10_pulse")
    comb_peak_slope = result_row(global_tests, "combined_29", "excel_peak_full", "dose_slope_log10_pulse")
    old_peak_high = result_row(global_tests, "old_17", "excel_peak_full", "high_pulses_mean_minus_1_pulse")
    comb_peak_high = result_row(global_tests, "combined_29", "excel_peak_full", "high_pulses_mean_minus_1_pulse")

    lines: list[str] = []
    lines.append("# DPM 功能成像左右偏侧化合并复检")
    lines.append("")
    lines.append("## 数据")
    lines.append("")
    lines.append("| batch | source | paired flies | sheets |")
    lines.append("|---|---|---:|---:|")
    for _, row in manifest.iterrows():
        lines.append(
            f"| {row['dataset']} | `{row['source_path']}` | {int(row['n_flies'])} | {int(row['n_sheets'])} |"
        )
    lines.append("")
    lines.append(
        "两个 Excel 都使用同一套窗口公式：1 pulse 为 `SUM(B37:B38)`，50 pulses 为 "
        "`SUM(B37:B53)`，100 pulses 为 `SUM(B37:B69)`，200 pulses 为 `SUM(B37:B103)`。"
    )
    lines.append("")
    lines.append("## 结论")
    lines.append("")
    lines.append(
        "旧批次原始数据已经在备份目录中找回，并与新批次一起按同一解析器重算。"
        "新数据单独看方向总体一致，但样本量为 12 且变异较大，所以对“刺激窗口求和”这个旧主指标没有明显提升显著性；"
        "合并后该指标仍显著，但 p 值基本维持在旧数据水平。"
    )
    lines.append("")
    lines.append(
        f"- 刺激窗口求和 log-dose slope：旧 17 只 mean={old_stim_slope['mean_right_minus_left']:.3f}, "
        f"paired t p={fmt_p(old_stim_slope['paired_t_p_two_sided'])}；新 12 只 mean="
        f"{new_stim_slope['mean_right_minus_left']:.3f}, p={fmt_p(new_stim_slope['paired_t_p_two_sided'])}；"
        f"合并 29 只 mean={comb_stim_slope['mean_right_minus_left']:.3f}, "
        f"p={fmt_p(comb_stim_slope['paired_t_p_two_sided'])}。"
    )
    lines.append(
        f"- peak log-dose slope：旧 17 只 mean={old_peak_slope['mean_right_minus_left']:.3f}, "
        f"p={fmt_p(old_peak_slope['paired_t_p_two_sided'])}；新 12 只 mean="
        f"{new_peak_slope['mean_right_minus_left']:.3f}, p={fmt_p(new_peak_slope['paired_t_p_two_sided'])}；"
        f"合并 29 只 mean={comb_peak_slope['mean_right_minus_left']:.3f}, "
        f"p={fmt_p(comb_peak_slope['paired_t_p_two_sided'])}。"
    )
    lines.append(
        f"- peak 的 high-pulse minus 1-pulse 对比从旧 p={fmt_p(old_peak_high['paired_t_p_two_sided'])} "
        f"提升到合并 p={fmt_p(comb_peak_high['paired_t_p_two_sided'])}，这是本次新增数据最明确增强的证据。"
    )
    lines.append("")
    lines.append(
        "因此，严谨表述应为：DPM 功能成像继续支持刺激依赖的左右偏侧化；新增数据主要加强了 "
        "peak / 高刺激响应随 pulse 增强的证据，而不是显著加强刺激窗口积分这个旧主指标。"
    )
    lines.append("")
    lines.append("## 全局比较")
    lines.append("")
    for metric, title in [("excel_stim_sum", "刺激窗口求和"), ("excel_peak_full", "全 trace peak")]:
        lines.append(f"### {title}")
        lines.append("")
        lines.append("| group | contrast | n | mean R-L | median R-L | t p | Wilcoxon p | sign p | right>left |")
        lines.append("|---|---|---:|---:|---:|---:|---:|---:|---:|")
        rows = global_tests[
            global_tests["metric"].eq(metric)
            & global_tests["contrast"].isin(
                ["all_pulses_mean", "high_pulses_mean_minus_1_pulse", "dose_slope_log10_pulse"]
            )
        ].sort_values(["dataset_group", "contrast"])
        for _, row in rows.iterrows():
            lines.append(
                f"| {row['dataset_group']} | {row['contrast']} | {int(row['n'])} | "
                f"{row['mean_right_minus_left']:.3f} | {row['median_right_minus_left']:.3f} | "
                f"{fmt_p(row['paired_t_p_two_sided'])} | {fmt_p(row['wilcoxon_p_two_sided'])} | "
                f"{fmt_p(row['sign_test_p_two_sided'])} | {int(row['positive_fly_count'])}/{int(row['n'])} |"
            )
        lines.append("")
    lines.append("## 单 pulse 检验")
    lines.append("")
    lines.append(
        "单 pulse 结果建议作为辅助读数。合并后，刺激窗口求和在 100 pulses 和 200 pulses 达到 "
        "paired t p<0.05；peak 在 100 pulses 和 200 pulses 的证据更强。"
    )
    lines.append("")
    lines.append("## 输出文件")
    lines.append("")
    lines.append("- `tables/dpm_data_manifest.csv`: 数据来源和样本量。")
    lines.append("- `tables/dpm_timecourse_long.csv`: 两批 Excel 解析后的长表。")
    lines.append("- `tables/dpm_side_metrics_long.csv`: 每只 fly、每侧 ROI 的指标。")
    lines.append("- `tables/dpm_per_pulse_tests.csv`: 单 pulse paired 左右差异。")
    lines.append("- `tables/dpm_global_contrasts.csv`: 全局 dose/peak/高刺激对比。")
    lines.append(f"- `{figure_path.relative_to(output_dir)}`: 合并复检摘要图。")
    lines.append(f"- `{paired_figure_path.relative_to(output_dir)}`: fly-level paired peak dose-slope 主读数图。")

    report = output_dir / "DPM_FUNCTIONAL_IMAGING_LATERALIZATION_20260603_CN.md"
    report.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report


def main() -> None:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    tables_dir = args.output_dir / "tables"
    figures_dir = args.output_dir / "figures"
    tables_dir.mkdir(parents=True, exist_ok=True)
    figures_dir.mkdir(parents=True, exist_ok=True)

    old_time, old_metrics, old_windows = load_excel_dataset(args.old_xlsx, "2026-05-21_old", "old")
    new_time, new_metrics, new_windows = load_excel_dataset(args.new_xlsx, "2026-06-03_new", "new")
    all_time = pd.concat([old_time, new_time], ignore_index=True)
    all_metrics = pd.concat([old_metrics, new_metrics], ignore_index=True)
    all_windows = pd.concat([old_windows, new_windows], ignore_index=True)

    manifest = pd.DataFrame(
        [
            {
                "dataset": "2026-05-21_old",
                "source_path": str(args.old_xlsx),
                "n_flies": old_metrics["fly"].nunique(),
                "n_sheets": old_metrics["sheet"].nunique(),
            },
            {
                "dataset": "2026-06-03_new",
                "source_path": str(args.new_xlsx),
                "n_flies": new_metrics["fly"].nunique(),
                "n_sheets": new_metrics["sheet"].nunique(),
            },
        ]
    )
    metric_names = ["excel_stim_sum", "excel_peak_full", "stim_window_sum_bc"]
    per_pulse, global_tests = build_test_tables(
        {
            "old_17": old_metrics,
            "new_12": new_metrics,
            "combined_29": all_metrics,
        },
        metric_names,
    )
    figure_path = plot_summary(per_pulse, global_tests, figures_dir)
    paired_figure_path = plot_paired_primary_endpoint(all_metrics, figures_dir)
    report = write_report(args.output_dir, manifest, all_metrics, per_pulse, global_tests, figure_path, paired_figure_path)

    manifest.to_csv(tables_dir / "dpm_data_manifest.csv", index=False)
    all_time.to_csv(tables_dir / "dpm_timecourse_long.csv", index=False)
    all_metrics.to_csv(tables_dir / "dpm_side_metrics_long.csv", index=False)
    all_windows.to_csv(tables_dir / "dpm_excel_response_windows.csv", index=False)
    per_pulse.to_csv(tables_dir / "dpm_per_pulse_tests.csv", index=False)
    global_tests.to_csv(tables_dir / "dpm_global_contrasts.csv", index=False)
    fly_level_dose_slopes(all_metrics, "excel_peak_full").to_csv(
        tables_dir / "dpm_fly_level_peak_dose_slopes.csv",
        index=False,
    )

    print(f"report: {report}")
    print(f"figure: {figure_path}")
    print(f"paired_figure: {paired_figure_path}")
    print(f"tables: {tables_dir}")


if __name__ == "__main__":
    main()
